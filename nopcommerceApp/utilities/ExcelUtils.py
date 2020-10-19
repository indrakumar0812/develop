import openpyxl

class HomePageData():

    test_data=[{'email': 'abc@gmail.com', 'pwd': 'test@1234', 'fname': 'Ram', 'lname': 'Kumar', 'gender': 'M', 'year': 1990, 'month': 'Jan', 'day': 17, 'cname': 'Testing', 'tax': 'Yes', 'newsletter': 'Test store 2', 'custroles': 'Registered', 'vendor': 1, 'active': 'Yes', 'admcomment': 'Hello world'}]


    def getTestData(sheetname):

        List =[]
        #workbook = openpyxl.load_workbook(".\\TestData\LoginData.xlsx")
        workbook = openpyxl.load_workbook("C:\\Users\\indrasen\\PycharmProjects\\nopcommerceApp\\TestData\\LoginData.xlsx")
        sheet = workbook[sheetname]
        for i in range(2, sheet.max_row+1):
            Dict = dict()
            for j in range(1, sheet.max_column+1):
                Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            List.append(Dict)
        #print(List)
        return List


#list.append(Dict)
#getTestData("customerDetails")
