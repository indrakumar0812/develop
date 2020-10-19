import openpyxl


class HomePageData:

    test_HomePage_data = [{"name":"Ram","email":"abc@gmail.com","gender":"Male"},{"name":"Simran","email":"xyz@gmail.com","gender":"Female"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = dict()
        workbook = openpyxl.load_workbook("C:\\Users\\indrasen\\Documents\\ExcelData\\PythonDemo.xlsx")
        sheet = workbook.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                # for j in range (1,sheet.max_column+1):
                for j in range(2, sheet.max_column + 1):
                    # print(sheet.cell(row=i, column=j).value)
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        #print(Dict)
        return[Dict]