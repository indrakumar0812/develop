import openpyxl

workbook = openpyxl.load_workbook("C:\\Users\\indrasen\\Documents\\ExcelData\\PythonDemo.xlsx")
sheet = workbook.active
row_count = sheet.max_row
col_count = sheet.max_column

for i in range (1,row_count+1):
    #print(sheet.cell(row=i,column=1).value)
    for j in range (1,col_count+1):
        print(sheet.cell(row=i,column=j).value)

